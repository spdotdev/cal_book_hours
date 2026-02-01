"""Excel XLSX file exporter using openpyxl."""

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from cal_exporter.models import CalendarEvent


class XLSXExporter:
    """Export events to Excel XLSX file format."""
    
    def __init__(self, output_path: str):
        """
        Initialize the XLSX exporter.
        
        Args:
            output_path: Path to the output XLSX file
        """
        self.output_path = Path(output_path)
    
    def export(self, events: List[CalendarEvent]) -> None:
        """
        Export events to an Excel file.
        
        Args:
            events: List of calendar events to export
        """
        # Ensure parent directory exists
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Calendar Events"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Define headers
        headers = [
            "Date",
            "Start Time",
            "End Time",
            "Duration (h)",
            "Duration",
            "Summary",
            "Description",
            "Location",
            "Hashtags",
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        for row, event in enumerate(events, 2):
            data = [
                event.start.strftime("%Y-%m-%d"),
                event.start.strftime("%H:%M"),
                event.end.strftime("%H:%M"),
                round(event.duration_hours, 2),
                event.duration_formatted,
                event.summary,
                event.description or "",
                event.location or "",
                ", ".join(event.hashtags),
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 4:  # Duration hours - right align numbers
                    cell.alignment = Alignment(horizontal="right")
        
        # Add summary row
        summary_row = len(events) + 3
        ws.cell(row=summary_row, column=1, value="Total Events:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(events))
        ws.cell(row=summary_row + 1, column=1, value="Total Hours:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=2, value=round(sum(e.duration_hours for e in events), 2))
        
        # Auto-adjust column widths
        column_widths = [12, 10, 10, 12, 10, 40, 50, 30, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save workbook
        wb.save(self.output_path)
